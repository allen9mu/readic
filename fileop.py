#-*- coding:utf-8 -*- 

import openpyxl,csv

#文件操作的内容。

import logging


def is_sheet_valid(wb, sheetname):
    return sheetname in wb.get_sheet_names()

def wirte_sheet(wb,points,SN,flag):
    logging.info('开始写入文件')
    if SN ==99:
        name = 'temp'
    else :
        name = SN
    
    if is_sheet_valid(wb,name):
        sheet = wb.get_sheet_by_name(name)
    else:
        sheet = wb.create_sheet(title=name)
        
    rows = sheet.max_row+1
    for key in points.keys():
        sheet.cell(row = rows, column=1).value= key
        sheet.cell(row = rows, column=2).value= points[key][0]                    
        sheet.cell(row = rows, column=3).value= points[key][1] 
        if flag==1:
            sheet.cell(row = rows, column=4).value= points[key][2]  
        rows =rows+1
    logging.info('写入文件结束'+str(len(points)))
    return wb


def wirte_csv(temp,OUTPUT_POS,IC_NUMBER):
    for file in temp.get_sheet_names():
        csvFile = open (OUTPUT_POS+'\\'+IC_NUMBER+'_'+file+'.csv','w',newline='')
        csvWrite = csv.writer(csvFile)
        sheet = temp.get_sheet_by_name(file)
        rows = sheet.max_row
        NO1ROW=0
        for row in range(2,rows+1):
            NN= sheet.cell(row = row,column=1).value
            TY= sheet.cell(row = row,column=2).value
            VA= sheet.cell(row = row,column=3).value
            
            if NO1ROW ==0 :
                csvWrite.writerow(['NodeNAME','Type','Value']) 
                csvWrite.writerow([NN,TY,VA])            
                NO1ROW =1
            else:
                csvWrite.writerow([NN,TY,VA]) 
