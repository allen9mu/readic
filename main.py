#-*- coding:utf-8 -*- 

import os,logging
import openpyxl
import time
from config import CONFIG,IOLIST_POS,IC_POS,IC_NUMBER,OUTPUT_POS
from ReadIC import get_ic_iolist,get_ic_keys,get_ic_SN,trans_ic_dict,get_ic_list
from ReadIOList import get_shns,get_io_prs,get_io_pl,get_io_spp
from fileop import wirte_sheet,wirte_csv

logging.info('开始计算')

#读取IOLIST
iolist = openpyxl.load_workbook(IOLIST_POS)

#建立中间文件
#中间文件 存储IOLIST生成的点表
temp_iolist = openpyxl.Workbook()
#中间文件 存储IC中导出的点表，准备生成IC导入文件
temp_ic = openpyxl.Workbook()

#读取sheet清单和点项关系清单
sheets = get_shns(CONFIG)
portrelates = get_io_prs(CONFIG)


#读取IOLIST生成点表清单
points_iolist={}
points_sum={}

for i in range(len(sheets)):
    name = sheets[i]
    relate = portrelates[name]
    pos = get_io_spp(iolist, name)
    points_iolist_m= get_io_pl(iolist, name, pos, portrelates)
    wirte_sheet(temp_iolist, points_iolist_m,99,1)
    points_iolist=dict(points_iolist,**points_iolist_m)
    print(points_iolist)
temp_iolist.save('test1.xlsx')    


#在配置文件中读取KEYS清单和点项关系清单
keys = get_ic_keys(CONFIG)


for file in os.listdir(IC_POS):
    SN= get_ic_SN(file)
    ic_dict = trans_ic_dict(IC_POS+"\\"+file)
    points_ic_m=get_ic_list(ic_dict, keys,points_iolist,SN)
    points_sum = dict(points_sum,**points_ic_m)
    if len(points_ic_m)>0:
        wirte_sheet(temp_ic, points_ic_m,SN,0)
temp_ic.save('test2.xlsx')  

wirte_csv(temp_ic, OUTPUT_POS, IC_NUMBER)

