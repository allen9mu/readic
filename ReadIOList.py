#!/usr/bin/env python  
#coding:utf-8  

import openpyxl 
import logging

#读取IOLIST，根据CONFIG生成点项清单

from config import CONFIG


#建立需在读取的sheet页的清单sheetnames shns
def get_shns(CONFIG):
    logging.info('开始生成sheetname清单')
    sheetnames = []
    for i in range(len(CONFIG)):
        if CONFIG[i][0]==2:
            sheetnames.append(CONFIG[i][1]) 
    logging.info('结束生成sheetname清单')
    return sheetnames


#生成IOLIST的点项对应关系清单portrelates prs
def get_io_prs(CNFIG):
    logging.info('开始生成portrelates清单')
    portrelates = {}
    for i in range(len(CONFIG)):
        list=[]
        if CONFIG[i][0]==2:
            for j in range(2,len(CONFIG[i])):
                list.append(CONFIG[i][j]) 
            portrelates[CONFIG[i][1]]=list        
    logging.info('结束生成portrelates清单')
    return  portrelates
    


#获取sheet页中SN和PN的位置sn pn pos spp
def get_io_spp(wb,sheetname):
    logging.info('开始生成SN和PN的位置')
    sheet_key = wb.get_sheet_by_name(sheetname)                            #激活相应的sheet
    columns = sheet_key.max_column       
    PN_pos=0                                                                #获得PN和SN的位置
    SN_pos=0
    for i in range(1,columns+1):
        search_sn_pn = sheet_key.cell(row = 1,column = i).value   
        if search_sn_pn == 'SN':
            SN_pos = i
            continue
        if search_sn_pn == 'PN':
            PN_pos = i
            continue
        if PN_pos != 0 & SN_pos != 0:                                       #找到PN和SN的位置后跳出循 
            break;
    pos = [SN_pos,PN_pos]  
    logging.info('结束生成SN和PN的位置')
    return pos
    
#生成目标点与参考点的对应清单io_pointlist pl
def get_io_pl(wb,sheetname,sn_pn_pos,prs):
    logging.info('生成目标点与参考点的对应清单 start')
    sheet_key = wb.get_sheet_by_name(sheetname)
    rows = sheet_key.max_row
    points = {}
    for i in range(3,rows+1):
        for j in range(len(prs[sheetname])):
            targetpoint = '.'+sheet_key.cell(row = i,column = sn_pn_pos[1]).value+'.'+ prs[sheetname][j][0]
            relatepoint = '.'+sheet_key.cell(row = i,column = sn_pn_pos[1]).value+'.'+ prs[sheetname][j][1]
            sn = sheet_key.cell(row = i,column = sn_pn_pos[0]).value
            points[targetpoint]= [relatepoint,sn,prs[sheetname][j][2]]
    logging.info('生成目标点与参考点的对应清单 end')    
    return points
        


if __name__ =='__main__':
    print(get_io_prs(CONFIG))
    
    
    